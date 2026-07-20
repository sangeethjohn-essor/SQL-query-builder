#!/usr/bin/env python3
"""Convert XLSX requirement sheets to structured Markdown."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Semantic column aliases (lowercase → canonical key)
HEADER_ALIASES: dict[str, str] = {
    "level": "level",
    "header/line": "level",
    "netsuite params": "target",
    "field": "target",
    "target": "target",
    "target field": "target",
    "column": "target",
    "required?": "required",
    "required": "required",
    "how to map": "rule_text",
    "logic": "rule_text",
    "mapping": "rule_text",
    "example": "example",
    "comments - francisco": "comments",
    "comments - jake": "comments_jake",
    "comments": "comments",
    "notes": "comments",
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def map_headers(raw_headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        key = normalize_header(header)
        if key in HEADER_ALIASES:
            mapping[idx] = HEADER_ALIASES[key]
    return mapping


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_required(value: str) -> bool:
    return value.lower() in {"yes", "y", "true", "1", "required"}


def extract_metadata(rows: list[list[str]]) -> dict[str, str]:
    """Scan first rows for source table hints in any cell."""
    metadata: dict[str, str] = {}
    blob = " ".join(" ".join(row) for row in rows[:3]).lower()
    if "outbound" in blob and "inbound" in blob:
        metadata["flow"] = "outbound_and_inbound"
    if "argents" in blob:
        metadata["three_pl"] = "argents"
    if "inv adjustments" in blob or "inventory adjustment" in blob:
        metadata["use_case"] = "inventory_adjustments"
    if "fact_finance_argents_outbound_shipment" in blob:
        metadata["source_outbound"] = "DWH.PROD.FACT_FINANCE_ARGENTS_OUTBOUND_SHIPMENT"
    if "fact_finance_argents_inbound_shipment" in blob:
        metadata["source_inbound"] = "DWH.PROD.FACT_FINANCE_ARGENTS_INBOUND_SHIPMENT"
    return metadata


def sheet_to_records(sheet) -> tuple[list[dict[str, str]], dict[str, str]]:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([cell_str(c) for c in row])

    # Drop fully empty rows
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return [], {}

    header_row_idx = 0
    col_map = map_headers(rows[header_row_idx])
    if "target" not in col_map.values():
        raise ValueError(f"Could not find target column header in sheet '{sheet.title}'")

    metadata = extract_metadata(rows)
    records: list[dict[str, str]] = []

    for row in rows[header_row_idx + 1 :]:
        if not any(row):
            continue
        record: dict[str, str] = {}
        for idx, key in col_map.items():
            if idx < len(row):
                record[key] = row[idx]
        target = record.get("target", "")
        if not target or target.lower() == "netsuite params":
            continue
        record["required_bool"] = "true" if parse_required(record.get("required", "")) else "false"
        records.append(record)

    return records, metadata


def records_to_markdown(
    source_path: Path,
    sheet_name: str,
    records: list[dict[str, str]],
    metadata: dict[str, str],
) -> str:
    lines = [
        f"# Requirements: {source_path.stem}",
        "",
        f"- **source_file:** `{source_path.name}`",
        f"- **sheet:** {sheet_name}",
    ]
    for k, v in metadata.items():
        lines.append(f"- **{k}:** {v}")
    lines.extend(["", "## Field mappings", ""])

    for rec in records:
        target = rec.get("target", "")
        level = rec.get("level", "")
        required = rec.get("required_bool", "false") == "true"
        rule_text = rec.get("rule_text", "").replace("\n", "  \n  ")
        example = rec.get("example", "").replace("\n", "  \n  ")
        comments = rec.get("comments", "")

        lines.append(f"### {target}")
        lines.append("")
        lines.append(f"- **level:** {level or 'Unknown'}")
        lines.append(f"- **required:** {'Yes' if required else 'No'}")
        if rule_text:
            lines.append(f"- **rule:** {rule_text}")
        if example:
            lines.append(f"- **example:** {example}")
        if comments:
            lines.append(f"- **comments:** {comments}")
        lines.append("")

    # Summary table for validate.py
    lines.extend(["## Required output columns", ""])
    lines.append("| Column | Level |")
    lines.append("|--------|-------|")
    for rec in records:
        if rec.get("required_bool") == "true":
            lines.append(f"| {rec.get('target', '')} | {rec.get('level', '')} |")
    lines.append("")

    return "\n".join(lines)


def convert_xlsx(xlsx_path: Path, output_path: Path | None = None) -> Path:
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    records, metadata = sheet_to_records(sheet)
    if not records:
        raise ValueError(f"No field mappings found in {xlsx_path}")

    md = records_to_markdown(xlsx_path, sheet.title, records, metadata)
    out = output_path or xlsx_path.with_suffix(".md")
    out.write_text(md, encoding="utf-8")
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert XLSX requirements to Markdown")
    parser.add_argument("xlsx", type=Path, help="Path to XLSX requirements file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output MD path (default: same name with .md extension)",
    )
    args = parser.parse_args(argv)

    if not args.xlsx.exists():
        print(f"Error: file not found: {args.xlsx}", file=sys.stderr)
        return 1

    try:
        out = convert_xlsx(args.xlsx, args.output)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
